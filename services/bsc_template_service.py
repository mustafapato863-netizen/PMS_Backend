from __future__ import annotations

import io
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from config.settings import DATA_DIR
from utils.performance_levels import normalize_performance_level


_STORE_PATH = Path(DATA_DIR) / "bsc_management_templates.json"
_TEMPLATE_PATH = Path(DATA_DIR) / "templates" / "Template_Managment.xlsx"
_PERSPECTIVE_ORDER = ["Financial", "Customer", "Internal Process", "Learning & Growth"]
_DEFAULT_STRATEGY_LINKS = [
    {"from": "Learning & Growth", "to": "Internal Process"},
    {"from": "Internal Process", "to": "Customer"},
    {"from": "Customer", "to": "Financial"},
]
_DEFAULT_PERSPECTIVE_META = {
    "Financial": {"label": "Financial", "focus": "Business profitability & revenue", "display_order": 1, "icon_key": "wallet"},
    "Customer": {"label": "Customer", "focus": "Stakeholder & patient experience", "display_order": 2, "icon_key": "users"},
    "Internal Process": {"label": "Internal Process", "focus": "Operational accuracy & compliance", "display_order": 3, "icon_key": "settings"},
    "Learning & Growth": {"label": "Learning & Growth", "focus": "Staff capacity & digital transformation", "display_order": 4, "icon_key": "graduation-cap"},
}


def _load_store() -> list[dict[str, Any]]:
    if not _STORE_PATH.exists():
        return []
    with _STORE_PATH.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _save_store(rows: list[dict[str, Any]]) -> None:
    _STORE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with _STORE_PATH.open("w", encoding="utf-8") as handle:
        json.dump(rows, handle, indent=2, ensure_ascii=False)


def _coerce_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1].strip()
    try:
        return float(text)
    except ValueError:
        return None


def _weight_to_fraction(value: Any) -> float:
    raw = _coerce_number(value)
    if raw is None:
        return 0.0
    return raw / 100.0 if raw > 1 else raw


def _slug(value: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return cleaned or "kpi"


def _parse_period(period: Any) -> tuple[str | None, int | None]:
    if period is None:
        return None, None
    if isinstance(period, datetime):
        return period.strftime("%B"), period.year
    if isinstance(period, date):
        return period.strftime("%B"), period.year
    if isinstance(period, (int, float)):
        try:
            parsed = from_excel(period)
            if isinstance(parsed, datetime):
                return parsed.strftime("%B"), parsed.year
            if isinstance(parsed, date):
                return parsed.strftime("%B"), parsed.year
        except Exception:
            pass
    text = str(period).strip()
    match = re.match(r"^([A-Za-z]+)\s+(\d{4})$", text)
    if match:
        return match.group(1), int(match.group(2))
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.strftime("%B"), parsed.year
        except ValueError:
            continue
    return None, None


def _direction_value(raw: Any) -> str:
    text = str(raw or "").strip().lower().replace(" ", "_")
    return "lower_better" if "lower" in text else "higher_better"


def _is_instruction_row(item: dict[str, Any]) -> bool:
    markers = {
        "Employee ID": {"keep employee code"},
        "Team": {"business unit / department"},
        "Employee Name": {"full employee name"},
        "Position": {"role / position name"},
        "Performance Level": {"managerial or corporate"},
        "Period": {"excel date, e.g. 5/1/2025"},
        "Perspective": {"financial / customer / internal process / learning & growth"},
        "KPI": {"kpi label"},
        "Direction": {"higher better or lower better"},
    }
    matches = 0
    for key, expected_values in markers.items():
        value = str(item.get(key) or "").strip().lower()
        if value in expected_values:
            matches += 1
    return matches >= 3


def _achievement_ratio(direction: str, actual_value: float | None, target_value: float | None) -> float | None:
    if actual_value is None or target_value is None:
        return None
    if direction == "lower_better":
        if actual_value == 0:
            return 1.0 if target_value == 0 else None
        return target_value / actual_value
    if target_value == 0:
        return 1.0 if actual_value == 0 else None
    return actual_value / target_value


class BSCTemplateService:
    sheet_name = "KPI's Data"

    def template_path(self) -> Path:
        return _TEMPLATE_PATH

    def parse_upload(self, contents: bytes) -> list[dict[str, Any]]:
        rows = self._read_template_rows(contents)
        if not rows:
            raise ValueError("No rows found in sheet KPI's Data")
        return rows

    def summarize_rows(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        levels = sorted({row["performance_level"] for row in rows})
        periods = sorted({f'{row["month"]} {row["year"]}' for row in rows if row.get("month") and row.get("year")})
        return {
            "rows_count": len(rows),
            "levels": levels,
            "periods": periods,
        }

    def build_database_payload(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        snapshots = []
        config_rows = []
        seen_employee_configs: set[tuple[Any, ...]] = set()
        snapshot_order = 0

        for row in rows:
            snapshot_order += 1
            kpi_key = _slug(str(row["kpi_label"]))
            snapshots.append({
                "employee_identifier": row["employee_id"],
                "employee_name": row["employee_name"],
                "position_name": row["position"],
                "performance_level": row["performance_level"],
                "month": row["month"],
                "year": int(row["year"]),
                "perspective_key": row["perspective"],
                "kpi_key": kpi_key,
                "kpi_label": row["kpi_label"],
                "actual_value": row.get("actual_value"),
                "display_order": snapshot_order,
            })

        grouped: dict[tuple[str, int, str, str, str], list[dict[str, Any]]] = {}
        for row in rows:
            kpi_key = _slug(str(row["kpi_label"]))
            key = (
                row["performance_level"],
                int(row["year"]),
                row["month"],
                row["position"],
                kpi_key,
            )
            grouped.setdefault(key, []).append(row)

        for (level, year, month, position, kpi_key), items in grouped.items():
            signatures: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
            for item in items:
                signature = (
                    item["perspective"],
                    item["direction"],
                    float(item["weight"]),
                    item.get("target_value"),
                    item.get("target_unit") or "%",
                    item["kpi_label"],
                )
                signatures.setdefault(signature, []).append(item)

            baseline_signature, baseline_items = max(
                signatures.items(),
                key=lambda entry: (len(entry[1]), sorted(row["employee_id"] for row in entry[1])[0]),
            )
            baseline = baseline_items[0]
            config_rows.append({
                "scope_type": "position",
                "position_name": position,
                "employee_identifier": None,
                "performance_level": level,
                "effective_month": month,
                "effective_year": year,
                "perspective_key": baseline_signature[0],
                "kpi_key": kpi_key,
                "kpi_label": baseline_signature[5],
                "direction": baseline_signature[1],
                "weight": baseline_signature[2],
                "target_value": baseline_signature[3],
                "target_unit": baseline_signature[4],
                "display_order": _PERSPECTIVE_ORDER.index(baseline_signature[0]) * 100 + len(config_rows) + 1,
            })

            for signature, signature_items in signatures.items():
                if signature == baseline_signature:
                    continue
                for item in signature_items:
                    employee_key = (level, year, month, item["employee_id"], kpi_key)
                    if employee_key in seen_employee_configs:
                        continue
                    seen_employee_configs.add(employee_key)
                    config_rows.append({
                        "scope_type": "employee",
                        "position_name": None,
                        "employee_identifier": item["employee_id"],
                        "performance_level": level,
                        "effective_month": month,
                        "effective_year": year,
                        "perspective_key": signature[0],
                        "kpi_key": kpi_key,
                        "kpi_label": signature[5],
                        "direction": signature[1],
                        "weight": signature[2],
                        "target_value": signature[3],
                        "target_unit": signature[4],
                        "display_order": _PERSPECTIVE_ORDER.index(signature[0]) * 100 + len(config_rows) + 1,
                    })

        return {
            "config_rows": config_rows,
            "snapshots": snapshots,
        }

    def save_upload(self, *, team: str, filename: str, contents: bytes, uploaded_by: str) -> dict[str, Any]:
        rows = self.parse_upload(contents)

        store = [item for item in _load_store() if item.get("team") != team]
        record = {
            "team": team,
            "filename": filename,
            "uploaded_by": uploaded_by,
            "uploaded_at": datetime.utcnow().isoformat(),
            "sheet_name": self.sheet_name,
            "rows": rows,
        }
        store.append(record)
        _save_store(store)

        return {
            "team": team,
            "filename": filename,
            "sheet_name": self.sheet_name,
            **self.summarize_rows(rows),
        }

    def get_dataset(self, *, team: str, performance_level: str, base_config: dict[str, Any]) -> dict[str, Any] | None:
        template = next((item for item in reversed(_load_store()) if item.get("team") == team), None)
        if not template:
            return None

        level = normalize_performance_level(performance_level)
        level_rows = [row for row in template.get("rows", []) if row.get("performance_level") == level]
        if not level_rows:
            return None

        records = self._rows_to_records(level_rows, team=team, performance_level=level)
        config = self._rows_to_config(level_rows, team=team, performance_level=level, base_config=base_config)
        return {
            "records": records,
            "config": config,
            "meta": {
                "filename": template.get("filename"),
                "uploaded_at": template.get("uploaded_at"),
                "uploaded_by": template.get("uploaded_by"),
            },
        }

    def _read_template_rows(self, contents: bytes) -> list[dict[str, Any]]:
        try:
            workbook = load_workbook(io.BytesIO(contents), data_only=True)
        except Exception as exc:
            raise ValueError("Only .xlsx template files are supported for Management Overview") from exc

        if self.sheet_name not in workbook.sheetnames:
            raise ValueError("Sheet KPI's Data was not found in the uploaded template")

        sheet = workbook[self.sheet_name]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]

        if not rows:
            return []

        headers = [str(value).strip() for value in rows[0]]
        required = {
            "Employee ID", "Team", "Employee Name", "Position", "Performance Level", "Period",
            "Perspective", "KPI", "Direction", "Weight", "Target Value", "Target Unit", "Actual Value",
        }
        missing = required - set(headers)
        if missing:
            raise ValueError(f"Template sheet KPI's Data is missing columns: {', '.join(sorted(missing))}")

        parsed_rows = []
        issues: list[str] = []
        for index, raw_row in enumerate(rows[1:], start=2):
            if not any(str(cell).strip() for cell in raw_row):
                continue
            item = dict(zip(headers, raw_row))
            if _is_instruction_row(item):
                continue
            month, year = _parse_period(item.get("Period"))
            level = normalize_performance_level(item.get("Performance Level"), allow_all=False)
            parsed = {
                "employee_id": str(item.get("Employee ID", "")).strip(),
                "team": str(item.get("Team", "")).strip(),
                "employee_name": str(item.get("Employee Name", "")).strip(),
                "position": str(item.get("Position", "")).strip(),
                "performance_level": level,
                "month": month,
                "year": year,
                "perspective": str(item.get("Perspective", "")).strip(),
                "kpi_label": str(item.get("KPI", "")).strip(),
                "direction": _direction_value(item.get("Direction")),
                "weight": _weight_to_fraction(item.get("Weight")),
                "target_value": _coerce_number(item.get("Target Value")),
                "target_unit": str(item.get("Target Unit", "")).strip() or "%",
                "actual_value": _coerce_number(item.get("Actual Value")),
            }
            if not parsed["employee_id"] or not parsed["employee_name"] or not parsed["position"] or not parsed["kpi_label"]:
                issues.append(f"Row {index} is missing employee or KPI identity fields")
                continue
            if not parsed["team"]:
                issues.append(f"Row {index} is missing Team")
                continue
            if parsed["perspective"] not in _PERSPECTIVE_ORDER:
                issues.append(f"Row {index} has invalid perspective '{parsed['perspective']}'")
                continue
            if not parsed["month"] or not parsed["year"]:
                issues.append(f"Row {index} has invalid Period '{item.get('Period')}'")
                continue
            if parsed["weight"] <= 0:
                issues.append(f"Row {index} has invalid Weight '{item.get('Weight')}'")
                continue
            parsed_rows.append(parsed)
        if issues:
            raise ValueError(issues[0])
        return parsed_rows

    def _rows_to_records(self, rows: list[dict[str, Any]], *, team: str, performance_level: str) -> list[dict[str, Any]]:
        grouped: dict[tuple[str, str, int], dict[str, Any]] = {}
        for row in rows:
            month = row.get("month")
            year = row.get("year")
            employee_id = row.get("employee_id")
            if not employee_id or not month or not year:
                continue

            key = (employee_id, month, int(year))
            record = grouped.setdefault(key, {
                "id": f"{employee_id}_{month}_{year}_{performance_level}",
                "employee_id": employee_id,
                "employee_name": row.get("employee_name"),
                "team": team,
                "month": month,
                "year": int(year),
                "performance_level": performance_level,
                "raw_data": {"Position": row.get("position"), "Period": f"{month} {year}"},
                "kpi_values": [],
                "evaluation": {"score": 0, "grade": "B"},
            })
            ratio = _achievement_ratio(row["direction"], row.get("actual_value"), row.get("target_value"))
            record["kpi_values"].append({
                "kpi_key": _slug(str(row["kpi_label"])),
                "actual_value": row.get("actual_value"),
                "target_value": row.get("target_value"),
                "achievement_ratio": ratio,
                "weight_applied": row.get("weight"),
                "contribution": row["weight"] * ratio if ratio is not None else None,
            })
        return list(grouped.values())

    def _rows_to_config(self, rows: list[dict[str, Any]], *, team: str, performance_level: str, base_config: dict[str, Any]) -> dict[str, Any]:
        thresholds = base_config.get("grade_thresholds", {"A": 90, "B": 80, "C": 70, "D": 60})
        base_bsc = base_config.get("balanced_scorecard", {}) or {}
        base_perspectives = {
            item.get("key"): item
            for item in base_bsc.get("perspectives", [])
            if item.get("key")
        }

        perspectives = []
        for key in _PERSPECTIVE_ORDER:
            meta = dict(_DEFAULT_PERSPECTIVE_META[key])
            meta.update(base_perspectives.get(key, {}))
            meta["key"] = key
            perspectives.append(meta)

        seen: set[str] = set()
        kpis = []
        for row in rows:
            kpi_key = _slug(str(row["kpi_label"]))
            if kpi_key in seen:
                continue
            seen.add(kpi_key)
            kpis.append({
                "key": kpi_key,
                "label": row["kpi_label"],
                "perspective": row["perspective"],
                "weight": row["weight"],
                "direction": row["direction"],
                "unit": row["target_unit"] or "%",
                "color": None,
                "rollup": "average",
            })

        return {
            "team": team,
            "performance_level": performance_level,
            "grade_thresholds": thresholds,
            "balanced_scorecard": {
                "enabled": True,
                "perspectives": perspectives,
                "strategy_map_links": base_bsc.get("strategy_map_links") or _DEFAULT_STRATEGY_LINKS,
            },
            "kpis": kpis,
        }


bsc_template_service = BSCTemplateService()
